#!/usr/bin/env python3
"""
Process the TC sheet in offense_codes_updated.xlsx by extracting statute text
from TN.doc files and populating columns E (elements) and F (statuteText).
"""

import openpyxl
from docx import Document
import os
import re
from collections import defaultdict

TN_DOC_FOLDER = 'TN.doc'
EXCEL_FILE = 'offense_codes_updated.xlsx'


def parse_citation(citation):
    """
    Parse a citation to extract chapter number and section number.
    
    Examples:
    - "545.001" -> chapter=545, section="545.001", subsection=None
    - "547.614(c)" -> chapter=547, section="547.614", subsection="(c)"
    - "550.021 3319" -> chapter=550, section="550.021", subsection=None
    - "521.342 (A)(2) 3241" -> chapter=521, section="521.342", subsection="(A)(2)"
    - "601.004(i)" -> chapter=601, section="601.004", subsection="(i)"
    """
    if not citation:
        return None, None, None
    
    citation = str(citation).strip()
    
    # Extract the main section part (before any space or with subsection)
    # Match patterns like: 545.001, 547.614(c), 521.342 (A)(2) 3241
    match = re.match(r'^(\d+)\.(\d+[A-Za-z]?)', citation)
    if not match:
        return None, None, None
    
    chapter = match.group(1)
    section = f"{match.group(1)}.{match.group(2)}"
    
    # Look for subsection in parentheses
    # Handle both "547.614(c)" and "521.342 (A)(2)"
    subsection_match = re.search(r'\(([a-zA-Z0-9\-]+)\)(?:\(([a-zA-Z0-9\-]+)\))?', citation)
    subsection = None
    if subsection_match:
        subsection = f"({subsection_match.group(1)})"
        if subsection_match.group(2):
            subsection += f"({subsection_match.group(2)})"
    
    return chapter, section, subsection


def get_tn_doc_path(chapter):
    """Find the TN.doc file for a given chapter."""
    if not chapter:
        return None
    
    # Try different filename patterns
    patterns = [
        f"tn.{chapter}.docx",
        f"TN.{chapter}.docx",
    ]
    
    for pattern in patterns:
        path = os.path.join(TN_DOC_FOLDER, pattern)
        if os.path.exists(path):
            return path
    
    # Case-insensitive search
    for filename in os.listdir(TN_DOC_FOLDER):
        if filename.lower() == f"tn.{chapter}.docx":
            return os.path.join(TN_DOC_FOLDER, filename)
    
    return None


def extract_section_from_doc(doc_path, section_number):
    """
    Extract a section from a TN.doc file.
    
    Returns (full_statute_text, all_subsection_texts) where:
    - full_statute_text: The complete section text with \n between paragraphs
    - all_subsection_texts: Dict mapping subsection letters to their text
    """
    doc = Document(doc_path)
    
    # Build section pattern to match "Sec. 545.001" etc.
    # Handle variations like "Sec. 545.001." or "Sec.545.001"
    section_pattern = rf"Sec\.\s*{re.escape(section_number)}\b"
    
    paragraphs_text = []
    in_section = False
    current_subsection = None
    subsection_texts = defaultdict(list)
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        
        # Check if this is the start of our section
        if re.match(section_pattern, text, re.IGNORECASE):
            in_section = True
            paragraphs_text.append(text)
            
            # Check for subsection in the same paragraph
            subsection_match = re.search(r'\([a-z]\)', text)
            if subsection_match:
                current_subsection = subsection_match.group()
                subsection_texts[current_subsection].append(text)
            continue
        
        # Check if we've hit the next section
        if in_section and re.match(r'Sec\.\s*\d+\.\d+', text, re.IGNORECASE):
            break
        
        # Check if we've hit the Acts section (legislative history)
        if in_section and re.match(r'^(Acts|Added by Acts|Amended by)', text, re.IGNORECASE):
            break
        
        if in_section:
            paragraphs_text.append(text)
            
            # Track subsections
            subsection_match = re.match(r'^(\([a-z](?:-\d+)?\))', text)
            if subsection_match:
                current_subsection = subsection_match.group(1)
                subsection_texts[current_subsection].append(text)
            elif current_subsection:
                subsection_texts[current_subsection].append(text)
    
    if not paragraphs_text:
        return None, {}
    
    # Join paragraphs with newlines
    full_text = '\n'.join(paragraphs_text)
    
    # Convert subsection dict values to joined text
    subsections = {k: '\n'.join(v) for k, v in subsection_texts.items()}
    
    return full_text, subsections


def remove_section_header(text):
    """
    Remove the section number and title from the beginning of text.
    E.g., "Sec. 550.026.  IMMEDIATE REPORT OF COLLISION.  (a)..." -> "(a)..."
    E.g., "Sec. 545.054.  PASSING TO THE LEFT:  SAFE DISTANCE.  (a)..." -> "(a)..."
    """
    if not text:
        return text
    
    # Pattern to match "Sec. XXX.XXX.  TITLE IN CAPS.  " at the start
    # Title can contain: uppercase letters, spaces, commas, dashes, semicolons, 
    # colons, slashes, apostrophes, parentheses, numbers
    # The title ends with a period followed by optional whitespace
    pattern = r'^Sec\.\s*\d+\.\d+[A-Za-z]?\.\s+[A-Z][A-Z0-9\s,\-;:/\'\(\)]+\.\s*'
    cleaned = re.sub(pattern, '', text, count=1)
    
    # If the first pattern didn't work, try a more aggressive approach
    if cleaned == text and re.match(r'^Sec\.', text):
        # Find where (a), (b), etc. starts and remove everything before it
        subsection_match = re.search(r'\s+(\([a-z]\)\s+)', text)
        if subsection_match:
            cleaned = text[subsection_match.start():].strip()
    
    return cleaned.strip()


def find_referenced_subsections(text):
    """
    Find all subsection references in text like "Subsection (a)" or "Subsection (b)(1)".
    Returns a list of referenced subsection identifiers like ['(a)', '(b)'].
    """
    # Match patterns like "Subsection (a)", "Subsection (a)(1)", "Subsections (a) and (b)"
    references = []
    
    # Pattern for "Subsection (x)" or "Subsection (x)(y)"
    pattern = r'[Ss]ubsection[s]?\s+\(([a-z](?:-\d+)?)\)(?:\s*(?:,|and|or)\s*\(([a-z](?:-\d+)?)\))?'
    matches = re.finditer(pattern, text)
    
    for match in matches:
        if match.group(1):
            references.append(f"({match.group(1)})")
        if match.group(2):
            references.append(f"({match.group(2)})")
    
    # Also handle "under Subsection (x)" pattern
    pattern2 = r'under\s+[Ss]ubsection\s+\(([a-z](?:-\d+)?)\)'
    for match in re.finditer(pattern2, text):
        ref = f"({match.group(1)})"
        if ref not in references:
            references.append(ref)
    
    return references


def extract_elements(full_text, subsection, subsection_texts):
    """
    Extract the elements (offense definition) for a specific subsection.
    
    Improved logic:
    1. Remove section headers (e.g., "Sec. 550.026. TITLE.")
    2. Find referenced subsections (e.g., "under Subsection (a)")
    3. Include referenced subsections FIRST, then the current subsection
    4. Focus on the specific offense elements
    """
    if not full_text:
        return None
    
    # Determine which subsection to look for
    target_subsection = subsection.lower() if subsection else None
    
    # Find the target subsection text
    target_text = None
    if target_subsection and subsection_texts:
        for key, text in subsection_texts.items():
            if key.lower() == target_subsection:
                target_text = text
                break
    
    # If no specific subsection found, try to find the one with "commits an offense"
    if not target_text:
        # Look for "commits an offense" pattern
        for sub_key in sorted(subsection_texts.keys()):
            sub_text = subsection_texts[sub_key]
            if re.search(r'commits\s+an\s+offense', sub_text, re.IGNORECASE):
                target_text = sub_text
                target_subsection = sub_key
                break
    
    # If still no target, use full text
    if not target_text:
        target_text = full_text
    
    # Find any referenced subsections in the target text
    referenced_subs = find_referenced_subsections(target_text)
    
    # Build the elements text
    elements_parts = []
    included_subs = set()
    
    # First, add any referenced subsections
    for ref_sub in referenced_subs:
        ref_sub_lower = ref_sub.lower()
        if ref_sub_lower not in included_subs:
            for key, text in subsection_texts.items():
                if key.lower() == ref_sub_lower:
                    # Remove section header from the referenced subsection
                    cleaned_text = remove_section_header(text)
                    elements_parts.append(cleaned_text)
                    included_subs.add(ref_sub_lower)
                    break
    
    # Then add the target subsection (if not already included)
    if target_subsection and target_subsection not in included_subs:
        cleaned_target = remove_section_header(target_text)
        elements_parts.append(cleaned_target)
    elif not target_subsection:
        # If no specific subsection, clean and add full text
        cleaned_target = remove_section_header(target_text)
        elements_parts.append(cleaned_target)
    
    # Join parts with newlines
    elements = '\n'.join(elements_parts) if elements_parts else remove_section_header(full_text)
    
    return elements


def main():
    print("=" * 60)
    print("PROCESSING TC SHEET")
    print("=" * 60)
    
    # Load the Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['TC']
    
    # Tracking stats
    total_rows = 0
    successful = 0
    failed = []
    no_doc_file = []
    
    # Cache for loaded documents to avoid reloading
    doc_cache = {}
    
    # Process each row (skip header)
    for row_num in range(2, ws.max_row + 1):
        citation = ws.cell(row=row_num, column=2).value  # Column B
        literal = ws.cell(row=row_num, column=1).value   # Column A (for debugging)
        
        if not citation:
            continue
        
        total_rows += 1
        
        # Parse the citation
        chapter, section, subsection = parse_citation(citation)
        
        if not chapter or not section:
            failed.append((row_num, citation, "Could not parse citation"))
            continue
        
        # Get the TN.doc file
        doc_path = get_tn_doc_path(chapter)
        
        if not doc_path:
            no_doc_file.append((row_num, citation, f"No TN.doc file for chapter {chapter}"))
            continue
        
        try:
            # Load document (use cache if available)
            if doc_path not in doc_cache:
                doc_cache[doc_path] = doc_path  # Store path for now
            
            # Extract section
            full_text, subsection_texts = extract_section_from_doc(doc_path, section)
            
            if not full_text:
                failed.append((row_num, citation, f"Section {section} not found in {doc_path}"))
                continue
            
            # Extract elements
            elements = extract_elements(full_text, subsection, subsection_texts)
            
            # Update the Excel cells
            ws.cell(row=row_num, column=5).value = elements      # Column E
            ws.cell(row=row_num, column=6).value = full_text     # Column F
            
            successful += 1
            
            # Print progress every 100 rows
            if successful % 100 == 0:
                print(f"Processed {successful} rows...")
                
        except Exception as e:
            failed.append((row_num, citation, str(e)))
    
    # Save the file
    print("\nSaving Excel file...")
    wb.save(EXCEL_FILE)
    wb.close()
    
    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total rows processed: {total_rows}")
    print(f"Successful: {successful}")
    print(f"Failed: {len(failed)}")
    print(f"No doc file: {len(no_doc_file)}")
    
    if no_doc_file:
        print("\n--- Missing TN.doc files (first 20) ---")
        for row, cit, reason in no_doc_file[:20]:
            print(f"  Row {row}: {cit} - {reason}")
    
    if failed:
        print("\n--- Failed extractions (first 20) ---")
        for row, cit, reason in failed[:20]:
            print(f"  Row {row}: {cit} - {reason}")
    
    # Show a sample of successful entries
    print("\n--- Sample of updated entries ---")
    ws = openpyxl.load_workbook(EXCEL_FILE)['TC']
    sample_count = 0
    for row_num in range(2, min(50, ws.max_row + 1)):
        statute_text = ws.cell(row=row_num, column=6).value
        if statute_text:
            citation = ws.cell(row=row_num, column=2).value
            print(f"\nRow {row_num} ({citation}):")
            print(f"  StatuteText (first 200 chars): {statute_text[:200]}...")
            sample_count += 1
            if sample_count >= 3:
                break


if __name__ == '__main__':
    main()
